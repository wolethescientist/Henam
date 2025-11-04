from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics import renderPDF
from io import BytesIO
from typing import List, Dict, Any
import logging
import os
from datetime import datetime
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import numpy as np

logger = logging.getLogger(__name__)


class ExportService:
    def __init__(self):
        self.styles = getSampleStyleSheet()
    
    def export_jobs_to_excel(self, jobs_data: List[Dict[str, Any]]) -> BytesIO:
        """Export jobs data to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs Report"
        
        # Headers
        headers = ["Job ID", "Title", "Status", "Priority", "Assigned To", "Created Date", "Due Date"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, job in enumerate(jobs_data, 2):
            ws.cell(row=row, column=1, value=job.get("id", ""))
            ws.cell(row=row, column=2, value=job.get("title", ""))
            ws.cell(row=row, column=3, value=job.get("status", ""))
            ws.cell(row=row, column=4, value=job.get("priority", ""))
            ws.cell(row=row, column=5, value=job.get("assigned_to", ""))
            ws.cell(row=row, column=6, value=job.get("created_at", ""))
            ws.cell(row=row, column=7, value=job.get("due_date", ""))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_finance_to_excel(self, finance_data: Dict[str, Any]) -> BytesIO:
        """Export finance data to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Finance Report"
        
        # Headers
        headers = ["Category", "Amount", "Type", "Date", "Description"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        transactions = finance_data.get("transactions", [])
        for row, transaction in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=transaction.get("category", ""))
            ws.cell(row=row, column=2, value=transaction.get("amount", 0))
            ws.cell(row=row, column=3, value=transaction.get("type", ""))
            ws.cell(row=row, column=4, value=transaction.get("date", ""))
            ws.cell(row=row, column=5, value=transaction.get("description", ""))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output    

    def export_jobs_to_pdf(self, jobs_data: List[Dict[str, Any]]) -> BytesIO:
        """Export jobs data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Jobs Report", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Jobs data table
        table_data = [["Job ID", "Title", "Status", "Priority", "Assigned To"]]
        
        for job in jobs_data:
            table_data.append([
                str(job.get("id", "")),
                job.get("title", ""),
                job.get("status", ""),
                job.get("priority", ""),
                job.get("assigned_to", "")
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Footer
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output
    
    def export_finance_to_pdf(self, finance_data: Dict[str, Any]) -> BytesIO:
        """Export finance data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Finance Report", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Finance data table
        table_data = [["Category", "Amount", "Type", "Date"]]
        
        transactions = finance_data.get("transactions", [])
        for transaction in transactions:
            table_data.append([
                transaction.get("category", ""),
                f"₦{transaction.get('amount', 0):,.2f}",
                transaction.get("type", ""),
                transaction.get("date", "")
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Footer
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output
    
    def export_invoice_to_pdf(self, invoice_data: Dict[str, Any]) -> BytesIO:
        """Export invoice data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Invoice", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Invoice details
        invoice_details = [
            ["Invoice ID", str(invoice_data.get("id", ""))],
            ["Job", invoice_data.get("job_title", "")],
            ["Client", invoice_data.get("client", "")],
            ["Amount", f"₦{invoice_data.get('amount', 0):,.2f}"],
            ["Paid Amount", f"₦{invoice_data.get('paid_amount', 0):,.2f}"],
            ["Pending Amount", f"₦{invoice_data.get('pending_amount', 0):,.2f}"],
            ["Due Date", str(invoice_data.get('due_date', ""))],
            ["Status", invoice_data.get('status', "")]
        ]
        
        # Create table
        table = Table(invoice_details)
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output
    
    def _create_chart_image(self, chart_type: str, data: Dict, title: str, width: int = 6, height: int = 4) -> BytesIO:
        """Create a chart image and return the BytesIO buffer."""
        try:
            try:
                plt.style.use('seaborn-v0_8')
            except:
                # Fallback if seaborn style is not available
                plt.style.use('default')
            
            fig, ax = plt.subplots(figsize=(width, height))
            
            # Set a professional color palette
            colors_palette = ['#2E86AB', '#A23B72', '#F18F01', '#C73E1D', '#6A994E', '#577590']
            
            if chart_type == 'pie':
                labels = list(data.keys())
                values = list(data.values())
                if values and sum(values) > 0:  # Only create pie chart if there's data
                    wedges, texts, autotexts = ax.pie(values, labels=labels, autopct='%1.1f%%', 
                                                    colors=colors_palette[:len(labels)], startangle=90)
                    # Make text more readable
                    for autotext in autotexts:
                        autotext.set_color('white')
                        autotext.set_fontweight('bold')
                else:
                    ax.text(0.5, 0.5, 'No data available', ha='center', va='center', transform=ax.transAxes)
                    
            elif chart_type == 'bar':
                labels = list(data.keys())
                values = list(data.values())
                if values and any(v > 0 for v in values):  # Only create bar chart if there's data
                    bars = ax.bar(labels, values, color=colors_palette[:len(labels)])
                    ax.set_ylabel('Amount (N)')
                    # Add value labels on bars
                    for bar in bars:
                        height = bar.get_height()
                        if height > 0:
                            ax.text(bar.get_x() + bar.get_width()/2., height,
                                   f'N{height:,.0f}', ha='center', va='bottom', fontweight='bold')
                    plt.xticks(rotation=45, ha='right')
                else:
                    ax.text(0.5, 0.5, 'No data available', ha='center', va='center', transform=ax.transAxes)
                
            elif chart_type == 'line':
                periods = list(data.keys())
                values = list(data.values())
                if values and periods:  # Only create line chart if there's data
                    ax.plot(periods, values, marker='o', linewidth=3, markersize=8, color=colors_palette[0])
                    ax.set_ylabel('Amount (N)')
                    ax.set_xlabel('Period')
                    plt.xticks(rotation=45, ha='right')
                    # Add grid for better readability
                    ax.grid(True, alpha=0.3)
                else:
                    ax.text(0.5, 0.5, 'No data available', ha='center', va='center', transform=ax.transAxes)
                
            ax.set_title(title, fontsize=14, fontweight='bold', pad=20)
            plt.tight_layout()
            
            # Save to BytesIO buffer instead of file
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight', facecolor='white')
            img_buffer.seek(0)
            plt.close()
            
            return img_buffer
            
        except Exception as e:
            logger.error(f"Error creating chart: {e}")
            plt.close()  # Make sure to close the figure even on error
            return None

    def export_financial_analytics_to_pdf(self, analytics_data: Dict[str, Any]) -> BytesIO:
        """Export financial analytics data to PDF format with beautiful charts and design."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4, 
                              rightMargin=50, leftMargin=50, 
                              topMargin=50, bottomMargin=50)
        story = []
        
        # Custom styles for a more beautiful design
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Title'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#2E86AB'),
            fontName='Helvetica-Bold',
            alignment=1  # Center alignment
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            spaceAfter=20,
            textColor=colors.HexColor('#577590'),
            fontName='Helvetica',
            alignment=1  # Center alignment
        )
        
        section_style = ParagraphStyle(
            'CustomSection',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=15,
            spaceBefore=25,
            textColor=colors.HexColor('#2E86AB'),
            fontName='Helvetica-Bold',
            borderWidth=2,
            borderColor=colors.HexColor('#2E86AB'),
            borderPadding=10,
            backColor=colors.HexColor('#F8F9FA')
        )
        
        # Header with company branding
        title = Paragraph("FINANCIAL ANALYTICS REPORT", title_style)
        story.append(title)
        
        # Report period with better styling
        period_text = f"Report Period: {analytics_data.get('period', 'N/A').title()}"
        start_date = analytics_data.get('start_date', '')
        end_date = analytics_data.get('end_date', '')
        if start_date and end_date:
            period_text += f" | {start_date} to {end_date}"
        
        period_para = Paragraph(period_text, subtitle_style)
        story.append(period_para)
        story.append(Spacer(1, 20))
        
        # Executive Summary with KPI cards
        exec_summary = Paragraph("EXECUTIVE SUMMARY", section_style)
        story.append(exec_summary)
        
        core_metrics = analytics_data.get('core_metrics', {})
        
        # Create KPI summary cards
        kpi_data = [
            ['Total Invoiced', f"N{core_metrics.get('total_invoiced', 0):,.2f}", 'Total Paid', f"N{core_metrics.get('total_paid', 0):,.2f}"],
            ['Total Pending', f"N{core_metrics.get('total_pending', 0):,.2f}", 'Total Expenses', f"N{analytics_data.get('total_expenses', 0):,.2f}"],
            ['Net Profit', f"N{core_metrics.get('net_profit', 0):,.2f}", 'Payment Rate', f"{core_metrics.get('payment_rate', 0):.1f}%"]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[120, 120, 120, 120])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8F9FA')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2E86AB')),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2E86AB')),
            ('ROUNDEDCORNERS', [5, 5, 5, 5])
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 30))
        
        # Revenue vs Expenses Chart
        revenue_expenses_title = Paragraph("REVENUE VS EXPENSES ANALYSIS", section_style)
        story.append(revenue_expenses_title)
        
        # Create revenue vs expenses chart
        revenue_expenses_data = {
            'Total Revenue': core_metrics.get('total_paid', 0),
            'Total Expenses': analytics_data.get('total_expenses', 0),
            'Net Profit': core_metrics.get('net_profit', 0)
        }
        
        try:
            chart_buffer = self._create_chart_image('bar', revenue_expenses_data, 
                                                  'Revenue vs Expenses Breakdown', 8, 5)
            if chart_buffer:
                chart_img = Image(chart_buffer, width=6*inch, height=3.5*inch)
                story.append(chart_img)
            else:
                story.append(Paragraph("Revenue chart could not be generated", self.styles['Normal']))
        except Exception as e:
            logger.error(f"Error creating revenue chart: {e}")
            story.append(Paragraph("Revenue chart could not be generated", self.styles['Normal']))
        
        story.append(Spacer(1, 30))
        
        # Invoice Status Distribution
        invoice_title = Paragraph("INVOICE STATUS DISTRIBUTION", section_style)
        story.append(invoice_title)
        
        invoice_distribution = analytics_data.get('invoice_status_distribution', [])
        if invoice_distribution:
            # Create pie chart for invoice status
            invoice_data = {}
            for status in invoice_distribution:
                invoice_data[status.get('status', 'Unknown')] = status.get('amount', 0)
            
            try:
                chart_buffer = self._create_chart_image('pie', invoice_data, 
                                                      'Invoice Status by Amount', 7, 5)
                if chart_buffer:
                    chart_img = Image(chart_buffer, width=5*inch, height=3.5*inch)
                    story.append(chart_img)
                else:
                    story.append(Paragraph("Invoice status chart could not be generated", self.styles['Normal']))
            except Exception as e:
                logger.error(f"Error creating invoice chart: {e}")
                story.append(Paragraph("Invoice status chart could not be generated", self.styles['Normal']))
            
            # Add detailed table
            invoice_status_data = [['Status', 'Count', 'Amount', 'Percentage']]
            total_amount = sum(status.get('amount', 0) for status in invoice_distribution)
            
            for status in invoice_distribution:
                amount = status.get('amount', 0)
                percentage = (amount / total_amount * 100) if total_amount > 0 else 0
                invoice_status_data.append([
                    status.get('status', ''),
                    str(status.get('count', 0)),
                    f"N{amount:,.2f}",
                    f"{percentage:.1f}%"
                ])
            
            invoice_table = Table(invoice_status_data, colWidths=[100, 80, 120, 80])
            invoice_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6'))
            ]))
            story.append(Spacer(1, 15))
            story.append(invoice_table)
        
        story.append(PageBreak())
        
        # Profit/Loss Trend Analysis
        trend_title = Paragraph("PROFIT/LOSS TREND ANALYSIS", section_style)
        story.append(trend_title)
        
        profit_loss_trend = analytics_data.get('profit_loss_trend', [])
        if profit_loss_trend:
            # Create line chart for trend
            trend_data = {}
            for trend in profit_loss_trend:
                period = trend.get('period', '')
                profit_loss = trend.get('profit_loss', 0)
                trend_data[period] = profit_loss
            
            try:
                chart_buffer = self._create_chart_image('line', trend_data, 
                                                      'Profit/Loss Trend Over Time', 8, 5)
                if chart_buffer:
                    chart_img = Image(chart_buffer, width=6*inch, height=3.5*inch)
                    story.append(chart_img)
                else:
                    story.append(Paragraph("Trend chart could not be generated", self.styles['Normal']))
            except Exception as e:
                logger.error(f"Error creating trend chart: {e}")
                story.append(Paragraph("Trend chart could not be generated", self.styles['Normal']))
            
            # Add detailed trend table
            trend_table_data = [['Period', 'Revenue', 'Expenses', 'Profit/Loss', 'Margin %']]
            for trend in profit_loss_trend:
                revenue = trend.get('revenue', 0)
                expenses = trend.get('expenses', 0)
                profit_loss = trend.get('profit_loss', 0)
                margin = (profit_loss / revenue * 100) if revenue > 0 else 0
                
                trend_table_data.append([
                    trend.get('period', ''),
                    f"N{revenue:,.2f}",
                    f"N{expenses:,.2f}",
                    f"N{profit_loss:,.2f}",
                    f"{margin:.1f}%"
                ])
            
            trend_table = Table(trend_table_data, colWidths=[80, 100, 100, 100, 80])
            trend_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86AB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6'))
            ]))
            story.append(Spacer(1, 15))
            story.append(trend_table)
        
        # Top Expense Categories
        story.append(Spacer(1, 30))
        expense_title = Paragraph("TOP EXPENSE CATEGORIES", section_style)
        story.append(expense_title)
        
        top_categories = analytics_data.get('top_expense_categories', [])
        if top_categories:
            # Create chart for top expenses
            expense_data = {}
            for category in top_categories[:5]:  # Top 5 categories
                expense_data[category.get('category', 'Unknown')] = category.get('total_amount', 0)
            
            try:
                chart_buffer = self._create_chart_image('bar', expense_data, 
                                                      'Top 5 Expense Categories', 8, 5)
                if chart_buffer:
                    chart_img = Image(chart_buffer, width=6*inch, height=3.5*inch)
                    story.append(chart_img)
                else:
                    story.append(Paragraph("Expense chart could not be generated", self.styles['Normal']))
            except Exception as e:
                logger.error(f"Error creating expense chart: {e}")
                story.append(Paragraph("Expense chart could not be generated", self.styles['Normal']))
        
        # Footer with insights
        story.append(Spacer(1, 40))
        footer_style = ParagraphStyle(
            'Footer',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#6C757D'),
            alignment=1,
            borderWidth=1,
            borderColor=colors.HexColor('#DEE2E6'),
            borderPadding=10,
            backColor=colors.HexColor('#F8F9FA')
        )
        
        footer_text = f"""
        <b>Report Summary:</b><br/>
        This comprehensive financial analytics report was generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}.<br/>
        The analysis includes revenue trends, expense breakdowns, and profitability metrics for informed decision-making.<br/>
        <br/>
        <b>Henam Facility Management Ltd</b> | Financial Analytics Dashboard
        """
        
        footer = Paragraph(footer_text, footer_style)
        story.append(footer)
        
        # Build PDF
        doc.build(story)
        output.seek(0)
        return output


# Global export service instance
export_service = ExportService()


def generate_invoice_pdf(invoice, db) -> BytesIO:
    """Generate PDF for a specific invoice and return the binary data."""
    # Import required modules
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Image, PageTemplate, Frame
    from reportlab.platypus.doctemplate import BaseDocTemplate
    from reportlab.pdfgen import canvas
    import os
    
    # Custom page template with watermark
    class WatermarkPageTemplate(PageTemplate):
        def __init__(self, id, frames, logo_path=None, **kwargs):
            PageTemplate.__init__(self, id, frames, **kwargs)
            self.logo_path = logo_path
            
        def beforeDrawPage(self, canvas, doc):
            """Add watermark before drawing page content"""
            if self.logo_path and os.path.exists(self.logo_path):
                try:
                    # Save current graphics state
                    canvas.saveState()
                    
                    # Set transparency (0.25 = 25% opacity, more visible overlay)
                    canvas.setFillAlpha(0.25)
                    canvas.setStrokeAlpha(0.25)
                    
                    # Get page dimensions
                    page_width, page_height = letter
                    
                    # Calculate position for watermark to overlay table area
                    # Make watermark larger and position it over the main content
                    watermark_width = 350
                    watermark_height = 120
                    x = (page_width - watermark_width) / 2
                    y = (page_height - watermark_height) / 2 + 50  # Shift up slightly to overlay table
                    
                    # Draw the watermark image
                    canvas.drawImage(self.logo_path, x, y, 
                                   width=watermark_width, 
                                   height=watermark_height,
                                   preserveAspectRatio=True,
                                   mask='auto')
                    
                    # Restore graphics state
                    canvas.restoreState()
                except Exception as e:
                    print(f"Warning: Could not add watermark: {e}")
    
    # Create PDF in memory with optimized margins for single page
    output = BytesIO()
    
    # Create custom document with watermark support
    doc = BaseDocTemplate(output, pagesize=letter, 
                         rightMargin=50, leftMargin=50, 
                         topMargin=50, bottomMargin=50)
    
    # Check if logo exists for watermark - try different formats
    logo_path = None
    logo_exists = False
    
    # Try different image formats
    for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
        test_logo_path = f"uploads/company_logo/henam_logo{ext}"
        if os.path.exists(test_logo_path):
            logo_path = test_logo_path
            logo_exists = True
            break
    
    # Create frame for content
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    
    # Create page template with watermark
    watermark_template = WatermarkPageTemplate(
        id='watermark',
        frames=[frame],
        logo_path=logo_path if logo_exists else None
    )
    
    # Add template to document
    doc.addPageTemplates([watermark_template])
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Clean professional styles with better fonts and spacing
    # Company header style - elegant and prominent
    company_style = ParagraphStyle(
        'CompanyHeader',
        parent=styles['Title'],
        fontSize=20,
        spaceAfter=10,
        spaceBefore=6,
        alignment=2,  # Right alignment
        textColor=colors.black,
        fontName='Helvetica-Bold',
        leading=24
    )
    
    # Invoice title style - elegant and eye-catching
    invoice_title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Heading1'],
        fontSize=28,
        spaceAfter=15,
        spaceBefore=10,
        alignment=2,  # Right alignment
        textColor=colors.HexColor('#34495E'),  # Elegant dark gray
        fontName='Helvetica-Bold',
        leading=32
    )
    
    # Company details style - clean and readable
    company_details_style = ParagraphStyle(
        'CompanyDetails',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=2,
        spaceBefore=1,
        alignment=2,  # Right alignment
        textColor=colors.black,
        fontName='Helvetica',
        leading=12
    )
    
    # Section header style - clean and clear
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=8,
        spaceBefore=12,
        textColor=colors.black,
        fontName='Helvetica-Bold',
        leading=16
    )
    
    # Body text style - clean and readable
    body_style = ParagraphStyle(
        'BodyText',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4,
        spaceBefore=2,
        textColor=colors.black,
        fontName='Helvetica',
        leading=13,
        alignment=0  # Left alignment
    )
    
    # Header section with small logo
    if logo_exists:
        try:
            # Add small header logo
            header_logo = Image(logo_path, width=120, height=40)
            elements.append(header_logo)
            elements.append(Spacer(1, 10))
        except Exception as e:
            print(f"Warning: Could not load header logo: {e}")
    
    # Add company information
    company_header = Paragraph("HENAM FACILITY MANAGEMENT LTD", company_style)
    elements.append(company_header)
    
    # Company details in a clean list format
    company_details = [
        "Dawaki, Abuja",
        "Abuja FCT",
        "NG",
        "09053121695",
        "henamcleaning@yahoo.com",
        "TIN: 31763224-0001 : RC: 7612266"
    ]
    
    for detail in company_details:
        detail_para = Paragraph(detail, company_details_style)
        elements.append(detail_para)
    
    # Add invoice number and date below company info
    invoice_number = Paragraph(f"Invoice #{invoice.invoice_number}", company_details_style)
    elements.append(invoice_number)
    
    invoice_date = Paragraph(f"Date {datetime.now().strftime('%d %b %Y')}", company_details_style)
    elements.append(invoice_date)
    
    # Add due date
    due_date = Paragraph(f"Due Date {invoice.due_date.strftime('%d %b %Y')}", company_details_style)
    elements.append(due_date)
    
    elements.append(Spacer(1, 18))  # Reduced space before next section
    
    # Get job and client information - handle both standalone and job-linked invoices
    from app.models import Job
    job = None
    if invoice.job_id:
        job = db.query(Job).filter(Job.id == invoice.job_id).first()
    
    # Bill To section - improved format
    bill_to_title = Paragraph("BILL TO", section_style)
    elements.append(bill_to_title)
    
    # Use client_name from invoice (works for both standalone and job-linked invoices)
    client_name = Paragraph(invoice.client_name, body_style)
    elements.append(client_name)
    elements.append(Spacer(1, 15))  # Reduced space
    
    # Itemized billing section - clean black format
    # Use NGN as currency symbol since Naira symbol has display issues
    naira_display = "NGN"
    # Determine service description - use job_type for standalone invoices or job title for linked invoices
    service_description = invoice.job_type if invoice.job_type else (job.title if job else "Professional Services")
    
    itemized_data = [
        ["Item", "Quantity", "Price", "Amount"],
        [service_description, "1", f"{naira_display}{invoice.amount:,.2f}", f"{naira_display}{invoice.amount:,.2f}"]
    ]
    
    # Create clean, professional itemized billing table
    itemized_table = Table(itemized_data, colWidths=[240, 80, 140, 140])
    itemized_table.setStyle(TableStyle([
        # Header row styling - elegant dark header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),  # Dark gray
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Data row styling - clean white background
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        
        # Alignment
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Item column left-aligned
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Quantity column centered
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),   # Price column right-aligned
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),   # Amount column right-aligned
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Padding for better spacing
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        
        # Clean borders
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#34495E')),  # Line under header
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#BDC3C7')),  # Bottom border
        ('LINEBEFORE', (0, 0), (0, -1), 1, colors.HexColor('#BDC3C7')),   # Left border
        ('LINEAFTER', (-1, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),  # Right border
        ('LINEAFTER', (0, 0), (0, -1), 1, colors.HexColor('#BDC3C7')),    # Column separators
        ('LINEAFTER', (1, 0), (1, -1), 1, colors.HexColor('#BDC3C7')),
        ('LINEAFTER', (2, 0), (2, -1), 1, colors.HexColor('#BDC3C7')),
    ]))
    
    elements.append(itemized_table)
    elements.append(Spacer(1, 20))
    
    # Description section - always show with improved styling
    description_title = Paragraph("DESCRIPTION:", section_style)
    elements.append(description_title)
    
    # Use job_details for standalone invoices, or invoice description, or default
    if invoice.job_details:
        description_text = invoice.job_details
    elif invoice.description:
        description_text = invoice.description
    else:
        description_text = "Professional services as agreed"
    
    description_para = Paragraph(description_text, body_style)
    elements.append(description_para)
    elements.append(Spacer(1, 15))  # Reduced space
    
    # Amount summary section - clean and simple
    amount_summary_data = [
        ["", f"Subtotal: {naira_display}{invoice.amount:,.2f}"],
        ["", f"Total: {naira_display}{invoice.amount:,.2f}"]
    ]
    
    # Create clean amount summary table
    amount_table = Table(amount_summary_data, colWidths=[350, 200])
    amount_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica'),
        ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 11),
        ('FONTSIZE', (1, 1), (1, 1), 13),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('LINEABOVE', (1, 1), (1, 1), 1, colors.HexColor('#BDC3C7')),  # Line above total
    ]))
    
    elements.append(amount_table)
    elements.append(Spacer(1, 18))
    
    # Payment Instructions section with improved styling
    payment_instructions_title = Paragraph("Payment Instructions", section_style)
    elements.append(payment_instructions_title)
    
    # Bank account details - simple list format with better styling
    bank_details = [
        "Henam Facility Management Limited",
        "Access Bank", 
        "1883625366",
        "Henam Cleaning Services",
        "Wema Bank",
        "0123104577"
    ]
    
    for detail in bank_details:
        elements.append(Paragraph(detail, body_style))
    
    elements.append(Spacer(1, 20))  # Reduced space before Amount due
    
    # Amount due section - professional and clear
    amount_due_data = [
        ["", f"Amount Due: {naira_display}{invoice.pending_amount:,.2f}"]
    ]
    
    # Create clean amount due table
    amount_due_table = Table(amount_due_data, colWidths=[350, 200])
    amount_due_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 16),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('LINEABOVE', (1, 0), (1, 0), 2, colors.black),  # Double line above for emphasis
        ('LINEBELOW', (1, 0), (1, 0), 2, colors.black),  # Double line below for emphasis
    ]))
    
    elements.append(amount_due_table)
    elements.append(Spacer(1, 15))  # Reduced space before footer
    
    # Footer - improved styling
    footer_text = (
        "Thank you for choosing Henam Facility Management Ltd!<br/>"
        "For any questions regarding this invoice, please contact us at henamcleaning@yahoo.com"
    )
    footer = Paragraph(footer_text, body_style)
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    # Reset BytesIO position to beginning
    output.seek(0)
    return output