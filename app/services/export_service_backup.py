from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from io import BytesIO
from typing import List, Dict, Any
import logging
import os
from datetime import datetime

logger = logging.getLogger(__name__)


class ExportService:
    def __init__(self):
        self.styles = getSampleStyleSheet()
    
    def export_jobs_to_excel(self, jobs_data: List[Dict[str, Any]]) -> BytesIO:
        """Export jobs data to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Jobs Report"
        
        # Headers
        headers = ["ID", "Title", "Client", "Start Date", "Progress", "Team", "Supervisor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, job in enumerate(jobs_data, 2):
            ws.cell(row=row, column=1, value=job.get("id", ""))
            ws.cell(row=row, column=2, value=job.get("title", ""))
            ws.cell(row=row, column=3, value=job.get("client", ""))
            ws.cell(row=row, column=4, value=str(job.get("start_date", "")))
            ws.cell(row=row, column=5, value=f"{job.get('progress', 0)}%")
            ws.cell(row=row, column=6, value=job.get("team_name", ""))
            ws.cell(row=row, column=7, value=job.get("supervisor_name", ""))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_finance_to_excel(self, finance_data: Dict[str, Any]) -> BytesIO:
        """Export finance data to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Finance Report"
        
        # Summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Invoiced", f"${finance_data.get('total_invoiced', 0):,.2f}"],
            ["Paid", f"${finance_data.get('paid', 0):,.2f}"],
            ["Pending", f"${finance_data.get('pending', 0):,.2f}"],
            ["Overdue Invoices", finance_data.get('overdue_invoices', 0)]
        ]
        
        for row, data in enumerate(summary_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 1:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def export_jobs_to_pdf(self, jobs_data: List[Dict[str, Any]]) -> BytesIO:
        """Export jobs data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Jobs Report", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Table data
        table_data = [["ID", "Title", "Client", "Start Date", "Progress", "Team", "Supervisor"]]
        
        for job in jobs_data:
            table_data.append([
                str(job.get("id", "")),
                job.get("title", ""),
                job.get("client", ""),
                str(job.get("start_date", "")),
                f"{job.get('progress', 0)}%",
                job.get("team_name", ""),
                job.get("supervisor_name", "")
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output
    
    def export_finance_to_pdf(self, finance_data: Dict[str, Any]) -> BytesIO:
        """Export finance data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Finance Report", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Summary data
        table_data = [
            ["Metric", "Value"],
            ["Total Invoiced", f"${finance_data.get('total_invoiced', 0):,.2f}"],
            ["Paid", f"${finance_data.get('paid', 0):,.2f}"],
            ["Pending", f"${finance_data.get('pending', 0):,.2f}"],
            ["Overdue Invoices", str(finance_data.get('overdue_invoices', 0))]
        ]
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output
    
    def export_invoice_to_pdf(self, invoice_data: Dict[str, Any]) -> BytesIO:
        """Export invoice data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Invoice", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Invoice details
        invoice_details = [
            ["Invoice ID", str(invoice_data.get("id", ""))],
            ["Job", invoice_data.get("job_title", "")],
            ["Client", invoice_data.get("client", "")],
            ["Amount", f"₦{invoice_data.get('amount', 0):,.2f}"],
            ["Paid Amount", f"₦{invoice_data.get('paid_amount', 0):,.2f}"],
            ["Pending Amount", f"₦{invoice_data.get('pending_amount', 0):,.2f}"],
            ["Due Date", str(invoice_data.get('due_date', ""))],
            ["Status", invoice_data.get('status', "")]
        ]
        
        # Create table
        table = Table(invoice_details)
        table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        output.seek(0)
        return output


# Global export service instance
export_service = ExportService()


def generate_invoice_pdf(invoice, db) -> BytesIO:
    """Generate PDF for a specific invoice and return the binary data."""
    # Import required modules
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import Image, PageTemplate, Frame
    from reportlab.platypus.doctemplate import BaseDocTemplate
    from reportlab.pdfgen import canvas
    import os
    
    # Custom page template with watermark
    class WatermarkPageTemplate(PageTemplate):
        def __init__(self, id, frames, logo_path=None, **kwargs):
            PageTemplate.__init__(self, id, frames, **kwargs)
            self.logo_path = logo_path
            
        def beforeDrawPage(self, canvas, doc):
            """Add watermark before drawing page content"""
            print(f"beforeDrawPage called with logo_path: {self.logo_path}")  # Debug
            if self.logo_path and os.path.exists(self.logo_path):
                print(f"Adding watermark from: {self.logo_path}")  # Debug
                try:
                    # Save current graphics state
                    canvas.saveState()
                    
                    # Set transparency (0.25 = 25% opacity, more visible overlay)
                    canvas.setFillAlpha(0.25)
                    canvas.setStrokeAlpha(0.25)
                    
                    # Get page dimensions
                    page_width, page_height = letter
                    
                    # Calculate position for watermark to overlay table area
                    # Make watermark larger and position it over the main content
                    watermark_width = 350
                    watermark_height = 120
                    x = (page_width - watermark_width) / 2
                    y = (page_height - watermark_height) / 2 + 50  # Shift up slightly to overlay table
                    
                    # Draw the watermark image
                    canvas.drawImage(self.logo_path, x, y, 
                                   width=watermark_width, 
                                   height=watermark_height,
                                   preserveAspectRatio=True,
                                   mask='auto')
                    
                    # Restore graphics state
                    canvas.restoreState()
                except Exception as e:
                    print(f"Warning: Could not add watermark: {e}")
    
    # Create PDF in memory with optimized margins for single page
    output = BytesIO()
    
    # Create custom document with watermark support
    doc = BaseDocTemplate(output, pagesize=letter, 
                         rightMargin=50, leftMargin=50, 
                         topMargin=50, bottomMargin=50)
    
    # Check if logo exists for watermark - try different formats
    logo_path = None
    logo_exists = False
    
    # Try different image formats
    for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp']:
        test_logo_path = f"uploads/company_logo/henam_logo{ext}"
        if os.path.exists(test_logo_path):
            logo_path = test_logo_path
            logo_exists = True
            print(f"Found logo at: {logo_path}")  # Debug print
            break
    
    if not logo_exists:
        print("No logo file found in uploads/company_logo/")  # Debug print
    
    # Create frame for content
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    
    # Create page template with watermark
    watermark_template = WatermarkPageTemplate(
        id='watermark',
        frames=[frame],
        logo_path=logo_path if logo_exists else None
    )
    
    # Add template to document
    doc.addPageTemplates([watermark_template])
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Clean professional styles with better fonts and spacing
    # Company header style - elegant and prominent
    company_style = ParagraphStyle(
        'CompanyHeader',
        parent=styles['Title'],
        fontSize=20,
        spaceAfter=10,
        spaceBefore=6,
        alignment=2,  # Right alignment
        textColor=colors.black,
        fontName='Helvetica-Bold',
        leading=24
    )
    
    # Invoice title style - elegant and eye-catching
    invoice_title_style = ParagraphStyle(
        'InvoiceTitle',
        parent=styles['Heading1'],
        fontSize=28,
        spaceAfter=15,
        spaceBefore=10,
        alignment=2,  # Right alignment
        textColor=colors.HexColor('#34495E'),  # Elegant dark gray
        fontName='Helvetica-Bold',
        leading=32
    )
    
    # Company details style - clean and readable
    company_details_style = ParagraphStyle(
        'CompanyDetails',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=2,
        spaceBefore=1,
        alignment=2,  # Right alignment
        textColor=colors.black,
        fontName='Helvetica',
        leading=12
    )
    
    # Section header style - clean and clear
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=8,
        spaceBefore=12,
        textColor=colors.black,
        fontName='Helvetica-Bold',
        leading=16
    )
    
    # Body text style - clean and readable
    body_style = ParagraphStyle(
        'BodyText',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=4,
        spaceBefore=2,
        textColor=colors.black,
        fontName='Helvetica',
        leading=13,
        alignment=0  # Left alignment
    )
    
    # Header section with small logo
    if logo_exists:
        try:
            print(f"Adding header logo from: {logo_path}")  # Debug
            # Add small header logo
            header_logo = Image(logo_path, width=120, height=40)
            elements.append(header_logo)
            elements.append(Spacer(1, 10))
            print("Header logo added successfully")  # Debug
        except Exception as e:
            print(f"Warning: Could not load header logo: {e}")
    else:
        print("No logo found for header")  # Debug
    
    # Add company information
    company_header = Paragraph("HENAM FACILITY MANAGEMENT LTD", company_style)
    elements.append(company_header)
    
    # Company details in a clean list format
    company_details = [
        "Dawaki, Abuja",
        "Abuja FCT",
        "NG",
        "09053121695",
        "henamcleaning@yahoo.com",
        "TIN: 31763224-0001 : RC: 7612266"
    ]
    
    for detail in company_details:
        detail_para = Paragraph(detail, company_details_style)
        elements.append(detail_para)
    
    # Add invoice number and date below company info
    invoice_number = Paragraph(f"Invoice #{invoice.invoice_number}", company_details_style)
    elements.append(invoice_number)
    
    invoice_date = Paragraph(f"Date {datetime.now().strftime('%d %b %Y')}", company_details_style)
    elements.append(invoice_date)
    
    # Add due date
    due_date = Paragraph(f"Due Date {invoice.due_date.strftime('%d %b %Y')}", company_details_style)
    elements.append(due_date)
    
    elements.append(Spacer(1, 18))  # Reduced space before next section
    
    # Get job and client information - handle both standalone and job-linked invoices
    from app.models import Job
    job = None
    if invoice.job_id:
        job = db.query(Job).filter(Job.id == invoice.job_id).first()
    
    # Bill To section - improved format
    bill_to_title = Paragraph("BILL TO", section_style)
    elements.append(bill_to_title)
    
    # Use client_name from invoice (works for both standalone and job-linked invoices)
    client_name = Paragraph(invoice.client_name, body_style)
    elements.append(client_name)
    elements.append(Spacer(1, 15))  # Reduced space
    
    # Itemized billing section - clean black format
    # Use NGN as currency symbol since Naira symbol has display issues
    naira_display = "NGN"
    # Determine service description - use job_type for standalone invoices or job title for linked invoices
    service_description = invoice.job_type if invoice.job_type else (job.title if job else "Professional Services")
    
    itemized_data = [
        ["Item", "Quantity", "Price", "Amount"],
        [service_description, "1", f"{naira_display}{invoice.amount:,.2f}", f"{naira_display}{invoice.amount:,.2f}"]
    ]
    
    # Create clean, professional itemized billing table
    itemized_table = Table(itemized_data, colWidths=[240, 80, 140, 140])
    itemized_table.setStyle(TableStyle([
        # Header row styling - elegant dark header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),  # Dark gray
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        
        # Data row styling - clean white background
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        
        # Alignment
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),    # Item column left-aligned
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),  # Quantity column centered
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),   # Price column right-aligned
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),   # Amount column right-aligned
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Padding for better spacing
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        
        # Clean borders
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#34495E')),  # Line under header
        ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#BDC3C7')),  # Bottom border
        ('LINEBEFORE', (0, 0), (0, -1), 1, colors.HexColor('#BDC3C7')),   # Left border
        ('LINEAFTER', (-1, 0), (-1, -1), 1, colors.HexColor('#BDC3C7')),  # Right border
        ('LINEAFTER', (0, 0), (0, -1), 1, colors.HexColor('#BDC3C7')),    # Column separators
        ('LINEAFTER', (1, 0), (1, -1), 1, colors.HexColor('#BDC3C7')),
        ('LINEAFTER', (2, 0), (2, -1), 1, colors.HexColor('#BDC3C7')),
    ]))
    
    elements.append(itemized_table)
    elements.append(Spacer(1, 20))
    
    # Description section - always show with improved styling
    description_title = Paragraph("DESCRIPTION:", section_style)
    elements.append(description_title)
    
    # Use job_details for standalone invoices, or invoice description, or default
    if invoice.job_details:
        description_text = invoice.job_details
    elif invoice.description:
        description_text = invoice.description
    else:
        description_text = "Professional services as agreed"
    
    description_para = Paragraph(description_text, body_style)
    elements.append(description_para)
    elements.append(Spacer(1, 15))  # Reduced space
    
    # Amount summary section - clean and simple
    amount_summary_data = [
        ["", f"Subtotal: {naira_display}{invoice.amount:,.2f}"],
        ["", f"Total: {naira_display}{invoice.amount:,.2f}"]
    ]
    
    # Create clean amount summary table
    amount_table = Table(amount_summary_data, colWidths=[350, 200])
    amount_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica'),
        ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 11),
        ('FONTSIZE', (1, 1), (1, 1), 13),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('LINEABOVE', (1, 1), (1, 1), 1, colors.HexColor('#BDC3C7')),  # Line above total
    ]))
    
    elements.append(amount_table)
    elements.append(Spacer(1, 18))
    
    # Payment Instructions section with improved styling
    payment_instructions_title = Paragraph("Payment Instructions", section_style)
    elements.append(payment_instructions_title)
    
    # Bank account details - simple list format with better styling
    bank_details = [
        "Henam Facility Management Limited",
        "Access Bank", 
        "1883625366",
        "Henam Cleaning Services",
        "Wema Bank",
        "0123104577"
    ]
    
    for detail in bank_details:
        elements.append(Paragraph(detail, body_style))
    
    elements.append(Spacer(1, 20))  # Reduced space before Amount due
    
    # Amount due section - professional and clear
    amount_due_data = [
        ["", f"Amount Due: {naira_display}{invoice.pending_amount:,.2f}"]
    ]
    
    # Create clean amount due table
    amount_due_table = Table(amount_due_data, colWidths=[350, 200])
    amount_due_table.setStyle(TableStyle([
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 16),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.black),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('LINEABOVE', (1, 0), (1, 0), 2, colors.black),  # Double line above for emphasis
        ('LINEBELOW', (1, 0), (1, 0), 2, colors.black),  # Double line below for emphasis
    ]))
    
    elements.append(amount_due_table)
    elements.append(Spacer(1, 15))  # Reduced space before footer
    
    # Footer - improved styling
    footer_text = (
        "Thank you for choosing Henam Facility Management Ltd!<br/>"
        "For any questions regarding this invoice, please contact us at henamcleaning@yahoo.com"
    )
    footer = Paragraph(footer_text, body_style)
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    
    # Reset BytesIO position to beginning
    output.seek(0)
    return output
    
    def export_personal_to_excel(self, personal_data: Dict[str, Any]) -> BytesIO:
        """Export personal data to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Personal Report"
        
        # Headers
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Personal data
        data_rows = [
            ["User Name", personal_data.get("user_name", "")],
            ["Email", personal_data.get("user_email", "")],
            ["Tasks Completed", personal_data.get("tasks_completed", 0)],
            ["Tasks Pending", personal_data.get("tasks_pending", 0)],
            ["Efficiency Score", f"{personal_data.get('efficiency_score', 0)}%"],
            ["Present Days", personal_data.get("present_days", 0)],
            ["Absent Days", personal_data.get("absent_days", 0)],
            ["Late Days", personal_data.get("late_days", 0)],
            ["Attendance Rate", f"{personal_data.get('attendance_percentage', 0)}%"]
        ]
        
        for row, (metric, value) in enumerate(data_rows, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream

    def export_personal_to_pdf(self, personal_data: Dict[str, Any]) -> BytesIO:
        """Export personal data to PDF format."""
        file_stream = BytesIO()
        doc = SimpleDocTemplate(file_stream, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Personal Performance Report", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # User info
        user_info = Paragraph(
            f"<b>User:</b> {personal_data.get('user_name', '')}<br/>"
            f"<b>Email:</b> {personal_data.get('user_email', '')}",
            self.styles['Normal']
        )
        elements.append(user_info)
        elements.append(Spacer(1, 20))
        
        # Task performance data
        task_data = [
            ["Metric", "Value"],
            ["Tasks Completed", str(personal_data.get("tasks_completed", 0))],
            ["Tasks Pending", str(personal_data.get("tasks_pending", 0))],
            ["Efficiency Score", f"{personal_data.get('efficiency_score', 0)}%"]
        ]
        
        task_table = Table(task_data)
        task_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(Paragraph("Task Performance", self.styles['Heading2']))
        elements.append(task_table)
        elements.append(Spacer(1, 20))
        
        # Attendance data
        attendance_data = [
            ["Metric", "Value"],
            ["Present Days", str(personal_data.get("present_days", 0))],
            ["Absent Days", str(personal_data.get("absent_days", 0))],
            ["Late Days", str(personal_data.get("late_days", 0))],
            ["Attendance Rate", f"{personal_data.get('attendance_percentage', 0)}%"]
        ]
        
        attendance_table = Table(attendance_data)
        attendance_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(Paragraph("Attendance Summary", self.styles['Heading2']))
        elements.append(attendance_table)
        elements.append(Spacer(1, 20))
        
        # Footer
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        file_stream.seek(0)
        
        return file_stream

    def export_teams_to_excel(self, teams_data: List[Dict[str, Any]]) -> BytesIO:
        """Export teams data to Excel format."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Teams Report"
        
        # Headers
        headers = ["Team ID", "Team Name", "Jobs Handled", "Avg Completion Time (days)", "Efficiency Score (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row, team in enumerate(teams_data, 2):
            ws.cell(row=row, column=1, value=team.get("team_id", ""))
            ws.cell(row=row, column=2, value=team.get("team_name", ""))
            ws.cell(row=row, column=3, value=team.get("jobs_handled", 0))
            ws.cell(row=row, column=4, value=team.get("average_completion_time", 0))
            ws.cell(row=row, column=5, value=f"{team.get('efficiency_score', 0)}%")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        return file_stream

    def export_financial_analytics_to_pdf(self, analytics_data: Dict[str, Any]) -> BytesIO:
        """Export financial analytics data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        
        # Title
        title_style = self.styles['Title']
        title = Paragraph("Financial Analytics Report", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Report period
        period_style = self.styles['Normal']
        period_text = f"Report Period: {analytics_data.get('period', 'N/A').title()}"
        period = Paragraph(period_text, period_style)
        story.append(period)
        
        start_date = analytics_data.get('start_date', '')
        end_date = analytics_data.get('end_date', '')
        if start_date and end_date:
            date_text = f"Date Range: {start_date} to {end_date}"
            date_para = Paragraph(date_text, period_style)
            story.append(date_para)
        
        story.append(Spacer(1, 20))
        
        # Revenue & Expense Summary
        summary_title = Paragraph("Revenue & Expense Summary", self.styles['Heading2'])
        story.append(summary_title)
        story.append(Spacer(1, 12))
        
        revenue_data = analytics_data.get('revenue_expense_summary', {})
        summary_data = [
            ['Metric', 'Amount'],
            ['Total Invoices Issued', f"₦{revenue_data.get('total_invoices_issued', 0):,.2f}"],
            ['Total Invoices Paid', f"₦{revenue_data.get('total_invoices_paid', 0):,.2f}"],
            ['Total Expenses', f"₦{revenue_data.get('total_expenses', 0):,.2f}"],
            ['Profit/Loss', f"₦{revenue_data.get('profit_loss', 0):,.2f}"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Top Expense Categories
        categories_title = Paragraph("Top Expense Categories", self.styles['Heading2'])
        story.append(categories_title)
        story.append(Spacer(1, 12))
        
        categories_data = [['Category', 'Total Amount', 'Count']]
        top_categories = analytics_data.get('top_expense_categories', [])
        for category in top_categories:
            categories_data.append([
                category.get('category', ''),
                f"₦{category.get('total_amount', 0):,.2f}",
                str(category.get('count', 0))
            ])
        
        categories_table = Table(categories_data)
        categories_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(categories_table)
        story.append(Spacer(1, 20))
        
        # Invoice Analytics
        invoice_title = Paragraph("Invoice Analytics", self.styles['Heading2'])
        story.append(invoice_title)
        story.append(Spacer(1, 12))
        
        invoice_data = analytics_data.get('invoice_analytics', {})
        invoice_summary_data = [
            ['Metric', 'Value'],
            ['Paid Invoices', str(invoice_data.get('paid_invoices', 0))],
            ['Unpaid Invoices', str(invoice_data.get('unpaid_invoices', 0))],
            ['Total Paid Amount', f"₦{invoice_data.get('total_paid_amount', 0):,.2f}"],
            ['Total Unpaid Amount', f"₦{invoice_data.get('total_unpaid_amount', 0):,.2f}"],
            ['Payment Rate', f"{invoice_data.get('paid_percentage', 0):.1f}%"]
        ]
        
        invoice_table = Table(invoice_summary_data)
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(invoice_table)
        story.append(Spacer(1, 20))
        
        # Profit/Loss Trend
        trend_title = Paragraph("Profit/Loss Trend", self.styles['Heading2'])
        story.append(trend_title)
        story.append(Spacer(1, 12))
        
        trend_data = [['Period', 'Revenue', 'Expenses', 'Profit/Loss']]
        profit_loss_trend = analytics_data.get('profit_loss_trend', [])
        for trend in profit_loss_trend:
            trend_data.append([
                trend.get('period', ''),
                f"₦{trend.get('revenue', 0):,.2f}",
                f"₦{trend.get('expenses', 0):,.2f}",
                f"₦{trend.get('profit_loss', 0):,.2f}"
            ])
        
        trend_table = Table(trend_data)
        trend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(trend_table)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        output.seek(0)
        return output
    
    def export_teams_to_pdf(self, teams_data: List[Dict[str, Any]]) -> BytesIO:
        """Export teams data to PDF format."""
        file_stream = BytesIO()
        doc = SimpleDocTemplate(file_stream, pagesize=letter)
        elements = []
        
        # Title
        title = Paragraph("Team Performance Report", self.styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Teams data table
        table_data = [["Team Name", "Jobs Handled", "Avg Completion Time", "Efficiency Score"]]
        
        for team in teams_data:
            table_data.append([
                team.get("team_name", ""),
                str(team.get("jobs_handled", 0)),
                f"{team.get('average_completion_time', 0)} days",
                f"{team.get('efficiency_score', 0)}%"
            ])
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Footer
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        elements.append(footer)
        
        # Build PDF
        doc.build(elements)
        file_stream.seek(0)
        
        return file_stream

    def export_financial_analytics_to_pdf(self, analytics_data: Dict[str, Any]) -> BytesIO:
        """Export financial analytics data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        
        # Title
        title_style = self.styles['Title']
        title = Paragraph("Financial Analytics Report", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Report period
        period_style = self.styles['Normal']
        period_text = f"Report Period: {analytics_data.get('period', 'N/A').title()}"
        period = Paragraph(period_text, period_style)
        story.append(period)
        
        start_date = analytics_data.get('start_date', '')
        end_date = analytics_data.get('end_date', '')
        if start_date and end_date:
            date_text = f"Date Range: {start_date} to {end_date}"
            date_para = Paragraph(date_text, period_style)
            story.append(date_para)
        
        story.append(Spacer(1, 20))
        
        # Revenue & Expense Summary
        summary_title = Paragraph("Revenue & Expense Summary", self.styles['Heading2'])
        story.append(summary_title)
        story.append(Spacer(1, 12))
        
        revenue_data = analytics_data.get('revenue_expense_summary', {})
        summary_data = [
            ['Metric', 'Amount'],
            ['Total Invoices Issued', f"₦{revenue_data.get('total_invoices_issued', 0):,.2f}"],
            ['Total Invoices Paid', f"₦{revenue_data.get('total_invoices_paid', 0):,.2f}"],
            ['Total Expenses', f"₦{revenue_data.get('total_expenses', 0):,.2f}"],
            ['Profit/Loss', f"₦{revenue_data.get('profit_loss', 0):,.2f}"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Top Expense Categories
        categories_title = Paragraph("Top Expense Categories", self.styles['Heading2'])
        story.append(categories_title)
        story.append(Spacer(1, 12))
        
        categories_data = [['Category', 'Total Amount', 'Count']]
        top_categories = analytics_data.get('top_expense_categories', [])
        for category in top_categories:
            categories_data.append([
                category.get('category', ''),
                f"₦{category.get('total_amount', 0):,.2f}",
                str(category.get('count', 0))
            ])
        
        categories_table = Table(categories_data)
        categories_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(categories_table)
        story.append(Spacer(1, 20))
        
        # Invoice Analytics
        invoice_title = Paragraph("Invoice Analytics", self.styles['Heading2'])
        story.append(invoice_title)
        story.append(Spacer(1, 12))
        
        invoice_data = analytics_data.get('invoice_analytics', {})
        invoice_summary_data = [
            ['Metric', 'Value'],
            ['Paid Invoices', str(invoice_data.get('paid_invoices', 0))],
            ['Unpaid Invoices', str(invoice_data.get('unpaid_invoices', 0))],
            ['Total Paid Amount', f"₦{invoice_data.get('total_paid_amount', 0):,.2f}"],
            ['Total Unpaid Amount', f"₦{invoice_data.get('total_unpaid_amount', 0):,.2f}"],
            ['Payment Rate', f"{invoice_data.get('paid_percentage', 0):.1f}%"]
        ]
        
        invoice_table = Table(invoice_summary_data)
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(invoice_table)
        story.append(Spacer(1, 20))
        
        # Profit/Loss Trend
        trend_title = Paragraph("Profit/Loss Trend", self.styles['Heading2'])
        story.append(trend_title)
        story.append(Spacer(1, 12))
        
        trend_data = [['Period', 'Revenue', 'Expenses', 'Profit/Loss']]
        profit_loss_trend = analytics_data.get('profit_loss_trend', [])
        for trend in profit_loss_trend:
            trend_data.append([
                trend.get('period', ''),
                f"₦{trend.get('revenue', 0):,.2f}",
                f"₦{trend.get('expenses', 0):,.2f}",
                f"₦{trend.get('profit_loss', 0):,.2f}"
            ])
        
        trend_table = Table(trend_data)
        trend_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(trend_table)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        output.seek(0)
        return output

    def export_financial_analytics_to_pdf(self, analytics_data: Dict[str, Any]) -> BytesIO:
        """Export financial analytics data to PDF format."""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter)
        story = []
        
        # Title
        title_style = self.styles['Title']
        title = Paragraph("Financial Analytics Report", title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Report period
        period_style = self.styles['Normal']
        period_text = f"Report Period: {analytics_data.get('period', 'N/A').title()}"
        period = Paragraph(period_text, period_style)
        story.append(period)
        
        start_date = analytics_data.get('start_date', '')
        end_date = analytics_data.get('end_date', '')
        if start_date and end_date:
            date_text = f"Date Range: {start_date} to {end_date}"
            date_para = Paragraph(date_text, period_style)
            story.append(date_para)
        
        story.append(Spacer(1, 20))
        
        # Financial Summary
        summary_title = Paragraph("Financial Summary", self.styles['Heading2'])
        story.append(summary_title)
        story.append(Spacer(1, 12))
        
        core_metrics = analytics_data.get('core_metrics', {})
        summary_data = [
            ['Metric', 'Amount'],
            ['Total Invoiced', f"₦{core_metrics.get('total_invoiced', 0):,.2f}"],
            ['Total Paid', f"₦{core_metrics.get('total_paid', 0):,.2f}"],
            ['Total Pending', f"₦{core_metrics.get('total_pending', 0):,.2f}"],
            ['Total Expenses', f"₦{analytics_data.get('total_expenses', 0):,.2f}"],
            ['Profit/Loss', f"₦{analytics_data.get('profit_loss', 0):,.2f}"],
            ['Payment Rate', f"{analytics_data.get('payment_rate', 0)}%"]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Top Expense Categories
        categories_title = Paragraph("Top Expense Categories", self.styles['Heading2'])
        story.append(categories_title)
        story.append(Spacer(1, 12))
        
        categories_data = [['Category', 'Total Amount', 'Count']]
        top_categories = analytics_data.get('top_expense_categories', [])
        for category in top_categories:
            categories_data.append([
                category.get('category', ''),
                f"₦{category.get('total_amount', 0):,.2f}",
                str(category.get('count', 0))
            ])
        
        if len(categories_data) > 1:  # Only show table if there's data
            categories_table = Table(categories_data)
            categories_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(categories_table)
        else:
            no_data = Paragraph("No expense data available for this period.", self.styles['Normal'])
            story.append(no_data)
        
        story.append(Spacer(1, 20))
        
        # Invoice Status Distribution
        invoice_title = Paragraph("Invoice Status Distribution", self.styles['Heading2'])
        story.append(invoice_title)
        story.append(Spacer(1, 12))
        
        invoice_status_data = [['Status', 'Count', 'Amount']]
        invoice_distribution = analytics_data.get('invoice_status_distribution', [])
        for status in invoice_distribution:
            invoice_status_data.append([
                status.get('status', ''),
                str(status.get('count', 0)),
                f"₦{status.get('amount', 0):,.2f}"
            ])
        
        if len(invoice_status_data) > 1:  # Only show table if there's data
            invoice_table = Table(invoice_status_data)
            invoice_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(invoice_table)
        else:
            no_data = Paragraph("No invoice data available.", self.styles['Normal'])
            story.append(no_data)
        
        story.append(Spacer(1, 20))
        
        # Profit/Loss Trend
        trend_title = Paragraph("Profit/Loss Trend", self.styles['Heading2'])
        story.append(trend_title)
        story.append(Spacer(1, 12))
        
        trend_data = [['Period', 'Revenue', 'Expenses', 'Profit/Loss']]
        profit_loss_trend = analytics_data.get('profit_loss_trend', [])
        for trend in profit_loss_trend:
            trend_data.append([
                trend.get('period', ''),
                f"₦{trend.get('revenue', 0):,.2f}",
                f"₦{trend.get('expenses', 0):,.2f}",
                f"₦{trend.get('profit_loss', 0):,.2f}"
            ])
        
        if len(trend_data) > 1:  # Only show table if there's data
            trend_table = Table(trend_data)
            trend_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(trend_table)
        else:
            no_data = Paragraph("No trend data available for this period.", self.styles['Normal'])
            story.append(no_data)
        
        # Footer
        story.append(Spacer(1, 30))
        footer_text = f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        footer = Paragraph(footer_text, self.styles['Normal'])
        story.append(footer)
        
        doc.build(story)
        output.seek(0)
        return output 
       return output